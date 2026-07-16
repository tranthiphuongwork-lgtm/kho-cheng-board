# -*- coding: utf-8 -*-
"""Doi chieu DOANH THU don chuyen khoan (Gobox) vs GIAO DICH TIEN VAO (Lark 'Sao ke').
TU CHUA (self-contained): chi phu thuoc gobox.py (qua gobox_orders) + openpyxl.
Da gop san cac ham Lark (token/list_records/send_card/upload_file) de KHONG can larkbase/config/messaging.

  1) Lay don Gobox trong ngay -> chia nhom PTTT (chuyen khoan / tien mat / khac), loc dung ngay.
  2) Lay giao dich 'in' tu bang 'Sao ke' (Lark).
  3) Khop don CHUYEN KHOAN <-> tien vao theo SO TIEN (uu tien ma don trong noi dung).
  4) Xuat Excel + gui thẻ tom tat + dinh kem file len Lark.

Chay:  python reconcile.py [start] [end] [--no-notify]
"""
import os, sys, json, uuid, datetime, urllib.request
import gobox_orders as GO

TZ = datetime.timezone(datetime.timedelta(hours=7))

# ---- Cau hinh Lark (env hoac mac dinh) ----
HOST      = os.getenv("LARK_HOST", "https://open.larksuite.com")
APP_ID    = os.getenv("LARK_APP_ID",  "cli_aa8d66518d619ed1")
APP_SEC   = os.getenv("LARK_APP_SEC") or os.getenv("LARK_APP_SECRET") or ""   # KHONG ghi cung: doc tu secret
BASE_APP  = os.getenv("LARK_BASE_APP", "SA7ebfOdLaUJ5fsVILIl1QGag7b")   # app_token cua Base
SAOKE_TABLE = os.getenv("GOBOX_SAOKE_TABLE", "tblQ4FASSV9Y6Ewl")
CONFIRM_CHAT = os.getenv("LARK_CONFIRM_CHAT_ID", "") or "oc_d284fd22a122a942ba6985414ecf0352"

F_SK_AMOUNT  = os.getenv("SAOKE_AMOUNT_FIELD", "số tiền")
F_SK_CONTENT = os.getenv("SAOKE_CONTENT_FIELD", "Nội dung")
F_SK_TYPE    = os.getenv("SAOKE_TYPE_FIELD", "Loại giao dịch")
F_SK_DATE    = os.getenv("SAOKE_DATE_FIELD", "Ngày giao dịch")
F_SK_ID      = os.getenv("SAOKE_ID_FIELD", "ID")
IN_VALUE     = os.getenv("SAOKE_IN_VALUE", "in")
TOLERANCE    = int(os.getenv("RECON_TOLERANCE", "0"))
IGFB_TOL     = int(os.getenv("RECON_IGFB_TOLERANCE", "999"))   # don ghi chu "igfb": cho lech < 1000d
CARD_MAX     = int(os.getenv("RECON_CARD_MAX", "20"))
UNVERIFIED_NOTE = os.getenv("RECON_UNVERIFIED_NOTE", "chưa nhận được tiền")
NOTIFY_DEFAULT = os.getenv("RECON_NOTIFY", "1") == "1"


# ================= LARK HELPERS (gop san) =================
def lark_token():
    r = urllib.request.Request(HOST + "/open-apis/auth/v3/tenant_access_token/internal",
        data=json.dumps({"app_id": APP_ID, "app_secret": APP_SEC}).encode(),
        headers={"Content-Type": "application/json"}, method="POST")
    return json.load(urllib.request.urlopen(r, timeout=30))["tenant_access_token"]

def gt(v):
    if isinstance(v, list): return "".join(x.get("text", "") for x in v if isinstance(x, dict))
    if isinstance(v, dict): return v.get("value") or v.get("text")
    return v

def lark_list(t, tid, fields):
    out, pt = [], None
    while True:
        url = HOST + f"/open-apis/bitable/v1/apps/{BASE_APP}/tables/{tid}/records/search?page_size=500" + (("&page_token=" + pt) if pt else "")
        r = urllib.request.Request(url, data=json.dumps({"field_names": fields}).encode(),
            headers={"Authorization": "Bearer " + t, "Content-Type": "application/json"}, method="POST")
        d = json.load(urllib.request.urlopen(r, timeout=60))["data"]; out += d.get("items", [])
        if d.get("has_more"): pt = d["page_token"]
        else: break
    return out

def lark_send_card(t, chat_id, card):
    body = {"receive_id": chat_id, "msg_type": "interactive", "content": json.dumps(card, ensure_ascii=False)}
    r = urllib.request.Request(HOST + "/open-apis/im/v1/messages?receive_id_type=chat_id",
        data=json.dumps(body).encode(), method="POST",
        headers={"Authorization": "Bearer " + t, "Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(r, timeout=30))

def lark_upload_file(t, path, fname):
    boundary = "----recon" + uuid.uuid4().hex
    with open(path, "rb") as f:
        content = f.read()
    parts = []
    def _add(name, value):
        parts.append(("--" + boundary).encode())
        parts.append(f'Content-Disposition: form-data; name="{name}"'.encode()); parts.append(b"")
        parts.append(str(value).encode())
    _add("file_type", "stream"); _add("file_name", fname)
    parts.append(("--" + boundary).encode())
    parts.append(f'Content-Disposition: form-data; name="file"; filename="{fname}"'.encode())
    parts.append(b"Content-Type: application/octet-stream"); parts.append(b""); parts.append(content)
    parts.append(("--" + boundary + "--").encode()); parts.append(b"")
    data = b"\r\n".join(parts)
    r = urllib.request.Request(HOST + "/open-apis/im/v1/files", data=data, method="POST",
        headers={"Authorization": "Bearer " + t, "Content-Type": "multipart/form-data; boundary=" + boundary})
    d = json.load(urllib.request.urlopen(r, timeout=60))
    return d.get("data", {}).get("file_key")

def lark_send_file(t, chat_id, file_key):
    body = {"receive_id": chat_id, "msg_type": "file", "content": json.dumps({"file_key": file_key})}
    r = urllib.request.Request(HOST + "/open-apis/im/v1/messages?receive_id_type=chat_id",
        data=json.dumps(body).encode(), method="POST",
        headers={"Authorization": "Bearer " + t, "Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(r, timeout=30))


# ================= SAO KE =================
def _to_date(ts_ms):
    try:
        return datetime.datetime.fromtimestamp(int(ts_ms) / 1000, TZ).date().isoformat()
    except Exception:
        return ""

def saoke_in(start_date, end_date):
    t = lark_token()
    rows = lark_list(t, SAOKE_TABLE, [F_SK_AMOUNT, F_SK_CONTENT, F_SK_TYPE, F_SK_DATE, F_SK_ID])
    out = []
    for it in rows:
        f = it["fields"]
        if gt(f.get(F_SK_TYPE)) != IN_VALUE:
            continue
        dd = _to_date(f.get(F_SK_DATE))
        if not (start_date <= dd <= end_date):
            continue
        amt = gt(f.get(F_SK_AMOUNT))
        try:
            amt = float(amt)
        except (TypeError, ValueError):
            continue
        out.append({"amount": amt, "content": gt(f.get(F_SK_CONTENT)) or "",
                    "id": gt(f.get(F_SK_ID)) or it.get("record_id"), "date": dd})
    return out


# ================= DOI CHIEU =================
def reconcile(orders, txns):
    txns = [dict(x, _used=False) for x in txns]
    matched, order_no_txn = [], []

    def _by_code(o):
        code = (o.get("code") or "").strip()
        if not code or len(code) < 4:
            return None
        for x in txns:
            if not x["_used"] and code in (x["content"] or ""):
                return x
        return None

    def _by_amount(o):
        cands = o.get("amounts") or ([o["amount"]] if o.get("amount") else [])
        tol = IGFB_TOL if o.get("igfb") else TOLERANCE
        best, bestkey = None, None
        for x in txns:
            if x["_used"]:
                continue
            diff = min(abs(x["amount"] - c) for c in cands)
            if diff > tol:
                continue
            key = (diff, 0 if x["date"] == o["date"] else 1)  # gan nhat, uu tien cung ngay
            if best is None or key < bestkey:
                best, bestkey = x, key
        return best

    for o in orders:
        x = _by_code(o)
        if x:
            x["_used"] = True; matched.append({"order": o, "txn": x, "by": "code"}); o["_done"] = True
    for o in orders:
        if o.get("_done"):
            continue
        x = _by_amount(o)
        if x:
            cands = o.get("amounts") or [o.get("amount")]
            diff = min(abs(x["amount"] - c) for c in cands if c is not None)
            x["_used"] = True
            matched.append({"order": o, "txn": x, "by": ("amount" if diff == 0 else f"igfb±{int(diff)}")})
        else:
            order_no_txn.append(o)
    txn_no_order = [x for x in txns if not x["_used"]]
    return matched, order_no_txn, txn_no_order


def run(start_date, end_date):
    g, oerr = GO.classify(start_date, end_date)
    if oerr:
        return {"ok": False, "stage": "gobox_orders", "err": oerr, "start": start_date, "end": end_date}
    orders = g["transfer"]
    txns = saoke_in(start_date, end_date)
    matched, order_no_txn, txn_no_order = reconcile(orders, txns)
    return {
        "ok": True, "start": start_date, "end": end_date,
        "n_orders": len(orders), "sum_orders": sum(o["amount"] or 0 for o in orders),
        "n_cash": len(g["cash"]), "sum_cash": sum(o["amount"] or 0 for o in g["cash"]),
        "n_other": len(g["other"]), "sum_other": sum(o["amount"] or 0 for o in g["other"]),
        "n_txn_in": len(txns), "sum_txn_in": sum(x["amount"] or 0 for x in txns),
        "n_matched": len(matched), "sum_matched": sum(m["order"]["amount"] or 0 for m in matched),
        "n_order_no_txn": len(order_no_txn), "sum_order_no_txn": sum(o["amount"] or 0 for o in order_no_txn),
        "n_txn_no_order": len(txn_no_order), "sum_txn_no_order": sum(x["amount"] or 0 for x in txn_no_order),
        "matched": matched, "order_no_txn": order_no_txn, "txn_no_order": txn_no_order,
        "cash_orders": g["cash"], "other_orders": g["other"],
    }


def _fmt(n):
    return f"{int(round(n or 0)):,}".replace(",", ".")

def summary_text(r):
    if not r.get("ok"):
        return "LỖI (" + str(r.get("stage", "?")) + "): " + str(r.get("err"))
    return "\n".join([
        f"Đối chiếu {r['start']}..{r['end']}",
        f"- Đơn chuyển khoản: {r['n_orders']} · {_fmt(r['sum_orders'])} đ",
        f"- Đơn tiền mặt: {r['n_cash']} · {_fmt(r['sum_cash'])} đ",
        f"- Tiền vào (Sao kê 'in'): {r['n_txn_in']} · {_fmt(r['sum_txn_in'])} đ",
        f"- KHỚP: {r['n_matched']} · {_fmt(r['sum_matched'])} đ",
        f"- Đơn CK chưa thấy tiền vào: {r['n_order_no_txn']} · {_fmt(r['sum_order_no_txn'])} đ",
        f"- Tiền vào chưa khớp đơn: {r['n_txn_no_order']} · {_fmt(r['sum_txn_no_order'])} đ",
    ])


def build_card(r):
    def _pct(a, b):
        return f" ({a/b*100:.0f}%)" if b else ""
    tmpl = "green" if r["n_order_no_txn"] == 0 else "orange"
    title = f"📊 Đối chiếu CK vs tiền vào · {r['start']}" + (f"..{r['end']}" if r['end'] != r['start'] else "")
    top = "\n".join([
        f"**Đơn chuyển khoản:** {r['n_orders']} · {_fmt(r['sum_orders'])} đ",
        f"💵 **Đơn tiền mặt:** {r['n_cash']} · {_fmt(r['sum_cash'])} đ",
        f"**Tiền vào (in):** {r['n_txn_in']} · {_fmt(r['sum_txn_in'])} đ",
        f"✅ **Khớp:** {r['n_matched']}{_pct(r['n_matched'], r['n_orders'])} · {_fmt(r['sum_matched'])} đ",
    ])
    def _blk(header, lines):
        shown = lines[:CARD_MAX]
        more = f"\n_… và {len(lines) - CARD_MAX} dòng khác (xem Excel)_" if len(lines) > CARD_MAX else ""
        return header + ("\n" + "\n".join(shown) if shown else "") + more
    def _oline(o):
        base = f"• `{o['code'] or '(không mã)'}` — **{_fmt(o['amount'])} đ**  _{o.get('date','')}_"
        note = (o.get('note') or '').strip()
        return base + (f"  · 📝 _{note}_" if note else "")
    ord_lines = [_oline(o) for o in r["order_no_txn"]]
    txn_lines = [f"• **{_fmt(x['amount'])} đ** — {(x['content'] or '')[:40]}  _{x.get('date','')}_" for x in r["txn_no_order"]]
    blk_ord = _blk(f"⚠️ **Đơn CK chưa thấy tiền vào: {r['n_order_no_txn']} · {_fmt(r['sum_order_no_txn'])} đ**", ord_lines)
    blk_txn = _blk(f"❔ **Tiền vào chưa khớp đơn: {r['n_txn_no_order']} · {_fmt(r['sum_txn_no_order'])} đ**", txn_lines)
    return {"config": {"wide_screen_mode": True},
            "header": {"template": tmpl, "title": {"tag": "plain_text", "content": title}},
            "elements": [
                {"tag": "div", "text": {"tag": "lark_md", "content": top}},
                {"tag": "hr"},
                {"tag": "div", "text": {"tag": "lark_md", "content": blk_ord}},
                {"tag": "hr"},
                {"tag": "div", "text": {"tag": "lark_md", "content": blk_txn}},
                {"tag": "note", "elements": [{"tag": "plain_text", "content": "File Excel chi tiết đính kèm bên dưới."}]},
            ]}


def write_xlsx(r, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(color="FFFFFF", bold=True); money = '#,##0'
    def _sheet(title, headers, rows, mcols=()):
        ws = wb.create_sheet(title); ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c); cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
        for row in rows: ws.append(row)
        for mc in mcols:
            for row in range(2, ws.max_row + 1): ws.cell(row, mc).number_format = money
        for col in range(1, len(headers) + 1):
            w = max([len(str(headers[col-1]))] + [len(str(ws.cell(rr, col).value or "")) for rr in range(2, ws.max_row + 1)])
            ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max(w + 2, 10), 48)
        ws.freeze_panes = "A2"
    ws = wb.active; ws.title = "Tổng hợp"
    ws["A1"] = "ĐỐI CHIẾU ĐƠN CHUYỂN KHOẢN vs TIỀN VÀO"; ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Kỳ: {r['start']} .. {r['end']}"
    ws.append([]); ws.append(["Nhóm", "Số đơn/GD", "Số tiền (đ)"])
    for c in range(1, 4):
        cell = ws.cell(4, c); cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
    for name, n, s in [
        ("Đơn chuyển khoản", r["n_orders"], r["sum_orders"]),
        ("Đơn tiền mặt", r["n_cash"], r["sum_cash"]),
        ("Đơn PTTT khác", r["n_other"], r["sum_other"]),
        ("Tiền vào (Sao kê 'in')", r["n_txn_in"], r["sum_txn_in"]),
        ("KHỚP", r["n_matched"], r["sum_matched"]),
        ("Đơn CK không thấy tiền vào", r["n_order_no_txn"], r["sum_order_no_txn"]),
        ("Tiền vào không khớp đơn", r["n_txn_no_order"], r["sum_txn_no_order"]),
    ]:
        ws.append([name, n, s])
    for row in range(5, ws.max_row + 1): ws.cell(row, 3).number_format = money
    ws.column_dimensions["A"].width = 32; ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 18
    _sheet("Khớp", ["Mã đơn", "Ngày đơn", "Doanh thu đơn (đ)", "Số tiền vào (đ)", "Nội dung tiền vào", "ID GD", "Khớp theo"],
           [[m["order"]["code"], m["order"]["date"], m["order"]["amount"], m["txn"]["amount"], m["txn"]["content"], m["txn"]["id"], m["by"]] for m in r["matched"]], mcols=(3, 4))
    _sheet("Đơn CK chưa có tiền vào", ["Mã đơn", "Ngày đơn", "Doanh thu đơn (đ)", "PTTT", "Ghi chú Gobox"],
           [[o["code"], o["date"], o["amount"], o.get("pay_txt", ""), (o.get("note") or "")] for o in r["order_no_txn"]], mcols=(3,))
    _sheet("Tiền vào chưa có đơn", ["Ngày", "Số tiền (đ)", "Nội dung", "ID GD"],
           [[x["date"], x["amount"], x["content"], x["id"]] for x in r["txn_no_order"]], mcols=(2,))
    _sheet("Đơn tiền mặt", ["Mã đơn", "Ngày đơn", "Số tiền (đ)", "PTTT"],
           [[o["code"], o["date"], o["amount"], o.get("pay_txt", "")] for o in r.get("cash_orders", [])], mcols=(3,))
    wb.save(path); return path


def notify(r, xlsx_path=None, chat_id=None):
    t = lark_token(); chat = chat_id or CONFIRM_CHAT; out = {"card": None, "file": None}
    try:
        out["card"] = lark_send_card(t, chat, build_card(r)).get("code")
    except Exception as e:
        out["card"] = "ERR " + str(e)
    if xlsx_path:
        try:
            fk = lark_upload_file(t, xlsx_path, os.path.basename(xlsx_path))
            if fk:
                out["file"] = lark_send_file(t, chat, fk).get("code")
        except Exception as e:
            out["file"] = "ERR " + str(e)
    return out


if __name__ == "__main__":
    flags = [a for a in sys.argv[1:] if a.startswith("-")]
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    start = args[0] if args else datetime.datetime.now(TZ).date().isoformat()
    end = args[1] if len(args) > 1 else start
    do_notify = NOTIFY_DEFAULT and "--no-notify" not in flags
    r = run(start, end)
    print(summary_text(r))
    if r.get("ok"):
        xlsx = f"reconcile_{start.replace('-','')}_{end.replace('-','')}.xlsx"
        try:
            write_xlsx(r, xlsx); print("Excel:", xlsx)
        except Exception as e:
            xlsx = None; print("Xuat Excel loi:", e)
        if do_notify:
            print("Lark:", json.dumps(notify(r, xlsx), ensure_ascii=False))
    slim = {k: v for k, v in r.items() if k not in ("matched", "order_no_txn", "txn_no_order", "cash_orders", "other_orders")}
    print(json.dumps(slim, ensure_ascii=False))
